#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ГО «Берегиня» — бекенд групи подовженого дня.

Що вміє:
  POST /api/zayavka   приймає анкету з form.html, пише в SQLite, шле лист адміну
  POST /api/login     вхід в адмінку за паролем
  GET  /admin         список заявок (потрібен вхід)
  GET  /admin/<id>    одна заявка (потрібен вхід)
  POST /admin/backup  копія бази: на сервер і, за вибором, на пошту (лише role=admin)

Запуск для розробки:   python3 app.py
Створити користувача:  python3 app.py adduser <логін> <пароль> [admin|teacher]
На сервері працює під gunicorn — див. deploy/beregynia.service
"""

import os, re, csv, json, sqlite3, smtplib, secrets, io, zipfile
from datetime import datetime, timedelta, timezone
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
from email.utils import formataddr

from flask import (Flask, request, session, redirect, url_for,
                   render_template, jsonify, abort, g)
from werkzeug.security import generate_password_hash, check_password_hash

# ============================================================
#  НАЛАШТУВАННЯ (з файла .env або зі змінних середовища)
# ============================================================
BASE = os.path.dirname(os.path.abspath(__file__))

def load_env():
    path = os.path.join(BASE, '.env')
    if not os.path.exists(path):
        return
    for line in open(path, encoding='utf-8'):
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        os.environ.setdefault(k.strip(), v.strip())

load_env()

DB_PATH     = os.environ.get('DB_PATH', os.path.join(BASE, 'beregynia.db'))
SECRET_KEY  = os.environ.get('SECRET_KEY', '')
ADMIN_MAIL  = os.environ.get('ADMIN_MAIL', '')        # кому слати нові заявки (можна кілька через кому)
SMTP_HOST   = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT   = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER   = os.environ.get('SMTP_USER', '')
SMTP_PASS   = os.environ.get('SMTP_PASS', '')
MAIL_FROM   = os.environ.get('MAIL_FROM', SMTP_USER)
SITE_URL    = os.environ.get('SITE_URL', 'https://iprostir.pp.ua')
# Копії бази лежать поруч із самою базою: цей каталог і так дозволений
# службі на запис (ReadWritePaths у beregynia.service), нічого налаштовувати
# додатково не треба.
BACKUP_DIR  = os.environ.get('BACKUP_DIR', os.path.join(os.path.dirname(DB_PATH), 'backups'))
BACKUP_KEEP = int(os.environ.get('BACKUP_KEEP', '10'))
PHONE_1     = os.environ.get('PHONE_1', '+380 97 382 33 79')
PHONE_2     = os.environ.get('PHONE_2', '+380 95 484 01 03')
DEV         = os.environ.get('DEV', '') == '1'

if not SECRET_KEY:
    if DEV:
        SECRET_KEY = 'dev-only-not-secret'
    else:
        raise SystemExit('Не задано SECRET_KEY. Створіть .env за зразком .env.example')

app = Flask(__name__)
app.config.update(
    SECRET_KEY=SECRET_KEY,
    SESSION_COOKIE_HTTPONLY=True,          # кука недоступна з JavaScript
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=not DEV,         # лише через HTTPS
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    MAX_CONTENT_LENGTH=256 * 1024,         # анкета не може важити більше 256 КБ
)

# ============================================================
#  БАЗА
# ============================================================
def db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
        # Якщо базу саме зараз хтось пише (бекап, ручний sqlite3), не падаємо
        # одразу, а чекаємо до 5 секунд. Інакше батько побачив би помилку
        # надсилання через те, що адміністратор відкрив базу в консолі.
        g.db.execute('PRAGMA busy_timeout = 5000')
        # Вбудований LIKE у SQLite опускає регістр лише для латиниці:
        # «шевченко» не знайшло б «Шевченка». Тому нижній регістр робить
        # Python — він знає і кирилицю.
        g.db.create_function('lower_uk', 1, lambda v: (v or '').lower())
    return g.db

@app.teardown_appcontext
def close_db(exc):
    conn = g.pop('db', None)
    if conn is not None:
        conn.close()

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(open(os.path.join(BASE, 'schema.sql'), encoding='utf-8').read())
    conn.commit()
    _migrate(conn)
    conn.close()

def _cleanup_orphans(conn):
    """Прибирає рядки дочірніх таблиць, у яких більше немає дитини.

    Каскадне видалення в SQLite працює лише коли для з'єднання ввімкнено
    PRAGMA foreign_keys. Консольний sqlite3 його НЕ вмикає, тому після
    ручного `DELETE FROM children` алергії та контакти лишаються висіти.
    Далі новій дитині може дістатися той самий id — і в її картці
    з'являться чужі медичні нотатки. Тому підчищаємо на кожному старті."""
    total = 0
    for table in ('sensitive', 'nmt', 'pickup_persons', 'schedule', 'attendance'):
        try:
            n = conn.execute(
                'DELETE FROM {} WHERE child_id NOT IN (SELECT id FROM children)'.format(table)
            ).rowcount
            total += max(n, 0)
        except sqlite3.OperationalError:
            pass       # таблиці ще немає
    if total:
        conn.commit()
        app.logger.warning('Прибрано %d осиротілих рядків у дочірніх таблицях', total)


def _migrate(conn):
    """Догоняє наявну базу до поточного набору полів.

    CREATE TABLE IF NOT EXISTS нову колонку в стару таблицю не додасть, тому
    після кожного розширення анкети треба ALTER TABLE. Тут це робиться саме,
    за списками колонок вище — вручну нічого писати не доведеться."""
    # Порожня таблиця → скидаємо лічильник, щоб наступна заявка була № 1.
    # Робимо це ТІЛЬКИ коли рядків немає: тоді номери ні з чим не зіткнуться.
    # Поки в базі є хоч одна заявка, номери не переспользовуються ніколи —
    # інакше лист «заявка № 5» через місяць вказував би на іншу дитину.
    _cleanup_orphans(conn)

    try:
        if conn.execute('SELECT COUNT(*) FROM children').fetchone()[0] == 0:
            n = conn.execute("DELETE FROM sqlite_sequence WHERE name='children'").rowcount
            if n:
                app.logger.warning('База порожня — нумерацію заявок скинуто, наступна буде № 1')
    except sqlite3.OperationalError:
        pass          # sqlite_sequence ще немає — база щойно створена

    for table, cols in (('children', CHILD_COLS + ['program', 'school_year'] + PUPIL_COLS),
                        ('sensitive', SENS_COLS), ('nmt', NMT_COLS),
                        ('pickup_persons', ['last', 'first', 'mid']),
                        ('login_attempts', ['login'])):
        have = {r[1] for r in conn.execute('PRAGMA table_info({})'.format(table))}
        if not have:                       # таблиці ще немає — її щойно створив скрипт
            continue
        for col in cols:
            if col not in have:
                conn.execute('ALTER TABLE {} ADD COLUMN {} TEXT'.format(table, col))
                app.logger.warning('База: додано колонку %s.%s', table, col)

    # Усі заявки, подані до появи напрямів, — це ГПД.
    # ALTER TABLE ADD COLUMN не проставляє DEFAULT наявним рядкам,
    # тому проставляємо самі, інакше вони випадуть з усіх фільтрів.
    try:
        n = conn.execute(
            "UPDATE children SET program='ГПД' WHERE program IS NULL OR program=''").rowcount
        if n:
            app.logger.warning('База: %d заявкам проставлено напрям «ГПД»', n)
        # індекси на щойно доданих колонках — тільки тут, коли колонки вже є
        conn.execute('CREATE INDEX IF NOT EXISTS idx_children_program ON children(program)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_children_year ON children(school_year)')
        # прізвище — те, за чим сортують усі списки; індекс саме тут,
        # бо на живій базі колонки ще немає (правило: індекси на нових
        # колонках лише в _migrate, ніколи в schema.sql)
        conn.execute('CREATE INDEX IF NOT EXISTS idx_children_last ON children(child_last)')
        # логін учня має бути унікальним, але порожніх — більшість,
        # тому індекс частковий: NULL і '' під обмеження не підпадають
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_children_login "
                     "ON children(login) WHERE login IS NOT NULL AND login <> ''")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_attempts_login ON login_attempts(login, ts)')
        # Статусів стало три. Старі не лишаємо висіти: їх не можна було б
        # ні поставити, ні прибрати, а у фільтрі й лічильниках вони б
        # плуталися. «Підтверджена» — це ще не зарахована, тож повертаємо
        # в «нову»; «відмова» — питання закрите, тобто «архів».
        for old, new in (('підтверджена', 'нова'), ('відмова', 'архів')):
            n3 = conn.execute('UPDATE children SET status=? WHERE status=?',
                              (new, old)).rowcount
            if n3:
                app.logger.warning('База: %d заявок зі статусу «%s» переведено в «%s»',
                                   n3, old, new)

        # заявки, подані до появи цієї колонки, — цьогорічні
        n2 = conn.execute("UPDATE children SET school_year=? WHERE school_year IS NULL OR school_year=''",
                          (school_year(),)).rowcount
        if n2:
            app.logger.warning('База: %d заявкам проставлено навчальний рік %s', n2, school_year())
    except sqlite3.OperationalError:
        pass

    conn.commit()

@app.template_filter('dmy')
def _dmy(v):
    """2016-05-14 → 14.05.2016. У базі дата лежить у машинному вигляді
    (щоб сортувалась), а на аркуші людина читає звичний порядок."""
    try:
        y, m, d = str(v).split('-')
        return '{}.{}.{}'.format(d, m, y)
    except (ValueError, AttributeError):
        return v or ''


def now():
    return datetime.now(timezone.utc).astimezone().strftime('%Y-%m-%d %H:%M:%S')

def client_ip():
    # за Caddy справжня адреса приходить у X-Forwarded-For
    fwd = request.headers.get('X-Forwarded-For', '')
    return (fwd.split(',')[0].strip() if fwd else request.remote_addr) or '?'

def log(action, child_id=None, who=None):
    db().execute(
        'INSERT INTO audit (ts, who, action, child_id, ip) VALUES (?,?,?,?,?)',
        (now(), who or session.get('login', '—'), action, child_id, client_ip()))
    db().commit()

# ============================================================
#  ПОЛЯ АНКЕТИ — один список, з якого все й будується
# ============================================================
FIELDS = [
    # (ключ у формі, підпис у листі й адмінці, чутливе?)
    ('child_last',     'Прізвище дитини',              False),
    ('child_first',    'Ім’я дитини',                  False),
    ('child_mid',      'По батькові дитини',           False),
    ('child_dob',      'Дата народження',              False),
    ('grade',          'Клас',                         False),
    ('school',         'Заклад освіти',                False),
    ('school_addr',    'Адреса закладу',               False),
    ('pickup_school',  'Забирати зі школи',            False),
    ('parent_last',    'Прізвище заявника',            False),
    ('parent_first',   'Ім’я заявника',                False),
    ('parent_mid',     'По батькові заявника',         False),
    ('parent_role',    'Ким доводиться дитині',        False),
    ('parent_dob',     'Дата народження заявника',     False),
    ('parent_phone',   'Телефон',                      False),
    ('parent_email',   'Пошта',                        False),
    ('student_phone',  'Телефон учня',                 False),
    ('student_email',  'Пошта учня',                   False),
    ('contact2_last',  'Прізвище другого контакту',    False),
    ('contact2_first', 'Ім’я другого контакту',        False),
    ('contact2_phone', 'Телефон другого контакту',     False),
    ('address',        'Адреса проживання',            False),
    ('pickup',         'Хто забирає дитину',           False),
    ('self_leave',     'Може йти додому сама',         False),
    ('self_time',      'Не раніше',                    False),
    ('has_allergy',    'Алергії',                      True),
    ('allergy_details','Деталі алергії',               True),
    ('meal_limits',    'Чого не їсть',                 True),
    ('health_notes',   'Що ще варто знати',            True),
    ('do_not_release', 'Кому не віддавати дитину',     True),
    ('expectations',   'Що важливо для батьків',       False),
    ('comment',        'Коментар',                     False),
    # --- довузівська підготовка (НМТ) ---
    ('career_help',    'Профорієнтаційна консультація', False),
    ('career_interest','Які професії цікавлять',       False),
    ('subjects',       'Предмети для підготовки',      False),
    ('needs',          'Чого очікує від навчання',     False),
    ('needs_other',    'Інше (потреби)',               False),
    ('level',          'Самооцінка рівня',             False),
    ('hard_topics',    'Що дається найважче',          False),
    ('format_pref',    'Бажаний формат занять',        False),
    ('time_pref',      'Бажаний час занять',           False),
    ('goal',           'Очікуваний результат',         False),
    ('speciality',     'Бажана спеціальність',         False),
    ('university',     'Заклад вищої освіти',          False),
    ('c_true',         'Згода: достовірність',         False),
    ('c_data',         'Згода: обробка даних',         False),
    ('c_health',       'Згода: дані про здоров’я',     False),
    ('c_medical',      'Згода: екстрена допомога',     False),
    ('c_photo',        'Згода: фото та відео',         False),
]
LABEL = {k: v for k, v, _ in FIELDS}

# ============================================================
#  НАПРЯМИ
#  ГПД — група подовженого дня, 1–8 клас.
#  НМТ — офлайн-підготовка до НМТ, 9–11 клас: інші поля,
#        інший перелік обов'язкових, окремі списки в адмінці.
#  Напрям НЕ приходить із форми, а виводиться з класу на сервері:
#  так його неможливо підмінити й неможливо забути передати.
# ============================================================
PROGRAMS = ('ГПД', 'НМТ')

# Статуси заявки. Три — і цього досить: заявка або щойно прийшла,
# або дитина ходить, або з нею вже все скінчено. Проміжні
# «підтверджена» й «відмова» лише плодили роботу з їх проставляння.
STATUSES = ('нова', 'зарахована', 'архів')

def school_year(d=None):
    """Навчальний рік у вигляді «2026/27».

    Рахуємо, а не зашиваємо: інакше щовересня хтось мав би правити код,
    і рано чи пізно не виправив би. Новий рік починається в серпні —
    тоді ж відкривається набір.
    """
    d = d or datetime.now()
    y = d.year - (1 if d.month < 8 else 0)
    return '{}/{}'.format(y, str(y + 1)[2:])


def program_for(grade):
    try:
        return 'НМТ' if int(grade) >= 9 else 'ГПД'
    except (TypeError, ValueError):
        return 'ГПД'

# Обов'язкові поля відрізняються: одинадцятикласник іде додому сам,
# вимагати від нього перелік осіб, які його забирають, безглуздо.
# Однаковий для обох напрямів: дані учня, один із батьків і все.
# Решту питаємо, але не тримаємо людину — перевантажена анкета
# просто не заповнюється до кінця.
#
# По батькові в обов'язкових НЕМАЄ навмисно: воно є не в кожного —
# і в дітей, народжених за кордоном, і просто за паспортом. Вимагати
# його означало б не пустити таку родину подати заявку взагалі.
REQUIRED_BY_PROGRAM = {
    'ГПД': ['child_last', 'child_first', 'child_dob', 'grade', 'school',
            'parent_last', 'parent_first', 'parent_role', 'parent_phone'],
    'НМТ': ['child_last', 'child_first', 'child_dob', 'grade', 'school',
            'parent_last', 'parent_first', 'parent_role', 'parent_phone'],
}

# child_name / parent_name / contact2_name у списку є, але з форми вони
# НЕ приходять — сервер збирає їх із складових у povne_imia(). Так само,
# як program виводиться з класу, а не приймається з анкети.
CHILD_COLS = ['child_name','child_last','child_first','child_mid',
              'child_dob','grade','school','school_addr','pickup_school',
              'parent_name','parent_last','parent_first','parent_mid',
              'parent_role','parent_dob','parent_phone','parent_email',
              'student_phone','student_email',
              'contact2_name','contact2_last','contact2_first','contact2_phone','address',
              'self_leave','self_time','expectations','comment',
              'c_true','c_data','c_health','c_medical','c_photo']

SENS_COLS  = ['has_allergy','allergy_details','meal_limits','health_notes','do_not_release']

# Доступ учня до навчальних матеріалів. Свідомо НЕ в CHILD_COLS: ті колонки
# заповнюються з анкети, а ці — тільки адміністратором. Якби вони потрапили
# до CHILD_COLS, форма могла б надіслати власний pass_hash.
PUPIL_COLS = ['login', 'pass_hash', 'pass_set_at', 'pass_by', 'last_seen']

# усе, що стосується лише довузівської підготовки — в окрему таблицю
NMT_COLS   = ['career_help','career_interest','subjects','needs','needs_other','level',
              'hard_topics','format_pref','time_pref','goal','speciality','university']

# застарілий загальний перелік лишається як запобіжник,
# фактично використовується REQUIRED_BY_PROGRAM
REQUIRED   = ['child_last', 'child_first', 'parent_last', 'parent_first', 'parent_phone']


def povne_imia(last, first, mid=''):
    """Збирає «Прізвище Ім'я По батькові», пропускаючи порожні частини.

    Єдине місце, де складається повне ім'я. Якщо збиратимеш його ще десь —
    рано чи пізно два місця розійдуться, і в листі буде одне, а в списку інше.
    Порядок саме такий: спершу прізвище, бо за ним шукають і сортують."""
    return ' '.join(x for x in (clean(last, 80), clean(first, 80), clean(mid, 80)) if x)

def clean(v, limit=2000):
    """Обрізаємо довжину й прибираємо керівні символи. Порожнє → ''."""
    if v is None:
        return ''
    v = str(v).replace('\x00', '').strip()
    return v[:limit]

# ============================================================
#  ПРИЙОМ АНКЕТИ
# ============================================================
@app.post('/api/zayavka')
def zayavka():
    f = request.form

    # --- антиспам ---
    if clean(f.get('website')):                       # honeypot: людина його не бачить
        return jsonify(ok=True, id=0)                 # боту показуємо «успіх» і мовчки викидаємо
    try:
        secs = int(f.get('fill_seconds') or 0)
    except ValueError:
        secs = 0
    if secs < 8:                                      # анкету неможливо заповнити за 8 секунд
        return jsonify(ok=False, error='too_fast'), 400

    # --- мінімальна перевірка ---
    data = {k: clean(f.get(k)) for k, _, _ in FIELDS}

    # Повні імена збирає сервер, з форми вони не приходять. Якби приходили,
    # можна було б надіслати прізвище окремо, а повне ім'я — чуже.
    data['child_name']    = povne_imia(data['child_last'], data['child_first'], data['child_mid'])
    data['parent_name']   = povne_imia(data['parent_last'], data['parent_first'], data['parent_mid'])
    data['contact2_name'] = povne_imia(data['contact2_last'], data['contact2_first'])

    program = program_for(data.get('grade'))
    for k in REQUIRED_BY_PROGRAM.get(program, REQUIRED):
        if not data[k]:
            return jsonify(ok=False, error='missing:' + k), 400
    if data['parent_email'] and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$', data['parent_email']):
        data['parent_email'] = ''                     # крива пошта не привід втрачати заявку

    conn = db()
    cur = conn.cursor()

    # Якщо заявок не лишилось (усе видалили), скидаємо лічильник — щоб наступна
    # була № 1, а не продовжувала стару нумерацію. Поки в базі є хоч один рядок,
    # номери не переспользовуються: інакше лист «заявка № 5» через місяць
    # указував би на іншу дитину.
    if cur.execute('SELECT COUNT(*) FROM children').fetchone()[0] == 0:
        cur.execute("DELETE FROM sqlite_sequence WHERE name='children'")

    # --- сира копія: комітимо ОКРЕМО й одразу.
    #     Навіть якщо далі щось піде не так, заявка не зникне безслідно. ---
    cur.execute('INSERT INTO raw_submissions (child_id, created_at, ip, payload) VALUES (?,?,?,?)',
                (None, now(), client_ip(), json.dumps(dict(f), ensure_ascii=False)))
    raw_id = cur.lastrowid
    conn.commit()

    try:
        # Порожня таблиця → скидаємо лічильник, щоб наступна заявка була № 1.
        # Поки в базі є хоч один рядок, номери не переспользовуються ніколи:
        # інакше лист «заявка № 5» через місяць указував би на іншу дитину.
        if cur.execute('SELECT COUNT(*) FROM children').fetchone()[0] == 0:
            _cleanup_orphans(conn)
            cur.execute("DELETE FROM sqlite_sequence WHERE name='children'")

        cols = ['created_at', 'source', 'fill_seconds', 'program', 'school_year'] + CHILD_COLS
        vals = [now(), 'site', secs, program, school_year()] + [data[c] for c in CHILD_COLS]
        cur.execute('INSERT INTO children ({}) VALUES ({})'.format(
            ','.join(cols), ','.join('?' * len(cols))), vals)
        child_id = cur.lastrowid

        # чутливе — в окрему таблицю
        cur.execute('INSERT INTO sensitive (child_id, {}) VALUES (?,{})'.format(
            ','.join(SENS_COLS), ','.join('?' * len(SENS_COLS))),
            [child_id] + [data[c] for c in SENS_COLS])

        # довузівська підготовка — теж окремо, і тільки для свого напряму
        if program == 'НМТ':
            cur.execute('INSERT INTO nmt (child_id, {}) VALUES (?,{})'.format(
                ','.join(NMT_COLS), ','.join('?' * len(NMT_COLS))),
                [child_id] + [data[c] for c in NMT_COLS])

        # Хто забирає: рядок «прізвище · ім'я · по батькові · телефон · ким доводиться | ...»
        # П'ять частин замість трьох — ці люди показують паспорт біля школи,
        # тож звіряти є з чим тільки коли прізвище лежить окремо.
        # Старий формат із трьох частин теж переживемо: чого бракує — те порожнє.
        for i, part in enumerate(p for p in data['pickup'].split('|') if p.strip()):
            bits = [b.strip() for b in part.split('·')]
            bits += [''] * (5 - len(bits))
            cur.execute('INSERT INTO pickup_persons '
                        '(child_id, ord, name, last, first, mid, phone, relation) '
                        'VALUES (?,?,?,?,?,?,?,?)',
                        (child_id, i + 1, povne_imia(bits[0], bits[1], bits[2]),
                         bits[0], bits[1], bits[2], bits[3], bits[4]))

        cur.execute('UPDATE raw_submissions SET child_id=? WHERE id=?', (child_id, raw_id))
        conn.commit()

    except Exception as e:
        # Відкат обов'язковий: інакше незавершена транзакція тримає базу
        # заблокованою, і наступні батьки взагалі не зможуть подати заявку.
        conn.rollback()
        app.logger.error('Заявку не збережено (сира копія № %s уціліла): %s', raw_id, e)
        return jsonify(ok=False, error='save_failed'), 500

    log('нова заявка', child_id, who='форма')

    # жоден лист не має права зламати прийом заявки — вона вже в базі
    try:
        send_admin_mail(child_id, data)
    except Exception as e:
        app.logger.error('Лист адміністратору не надіслано: %s', e)

    if data.get('parent_email'):
        try:
            send_parent_mail(child_id, data)
        except Exception as e:
            app.logger.error('Лист батькам не надіслано: %s', e)

    return jsonify(ok=True, id=child_id)

# ============================================================
#  ЛИСТ АДМІНІСТРАТОРУ
# ============================================================
def send_admin_mail(child_id, data):
    if not (SMTP_USER and SMTP_PASS and ADMIN_MAIL):
        app.logger.warning('Пошта не налаштована — лист пропущено')
        return

    rows = []
    for key, label, _sens in FIELDS:
        val = data.get(key, '')
        if not val:
            continue      # порожні поля іншого напряму просто не показуємо
        rows.append(
            '<tr><td style="padding:7px 12px;border-bottom:1px solid #E4E7EC;'
            'color:#475467;font-size:14px;white-space:nowrap;vertical-align:top">{}</td>'
            '<td style="padding:7px 12px;border-bottom:1px solid #E4E7EC;'
            'color:#101828;font-size:15px">{}</td></tr>'.format(
                esc(label), esc(val).replace('\n', '<br>')))

    html = """<!doctype html><html><body style="margin:0;background:#F5F6F7;
padding:24px;font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif">
<div style="max-width:640px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;
border:1px solid #E4E7EC">
  <div style="background:#2F6F4E;color:#fff;padding:20px 24px">
    <div style="font-size:13px;opacity:.85;letter-spacing:.04em;text-transform:uppercase">Нова заявка № {id}</div>
    <div style="font-size:21px;font-weight:700;margin-top:4px">{name}</div>
  </div>
  <table style="width:100%;border-collapse:collapse">{rows}</table>
  <div style="padding:18px 24px;background:#FAFAF7;border-top:1px solid #E4E7EC">
    <a href="{url}/admin/{id}" style="display:inline-block;background:#2F6F4E;color:#fff;
       text-decoration:none;padding:11px 20px;border-radius:8px;font-weight:600;font-size:15px">
       Відкрити в адмінці</a>
  </div>
</div>
<div style="max-width:640px;margin:14px auto 0;color:#98A2B3;font-size:12px;text-align:center">
  Лист містить персональні дані дитини. Не пересилайте його далі.
</div>
</body></html>""".format(id=child_id, name=esc(data['child_name']),
                         rows=''.join(rows), url=SITE_URL)

    msg = MIMEText(html, 'html', 'utf-8')
    msg['Subject'] = Header('Нова заявка № {} — {}'.format(child_id, data['child_name']), 'utf-8')
    msg['From']    = formataddr((str(Header('ГПД «Берегиня»', 'utf-8')), MAIL_FROM))
    msg['To']      = ADMIN_MAIL

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.sendmail(MAIL_FROM, [a.strip() for a in ADMIN_MAIL.split(',')], msg.as_string())

# ============================================================
#  ЛИСТ-ПІДТВЕРДЖЕННЯ БАТЬКАМ
#  Надсилається, лише якщо батько лишив пошту — вона необов'язкова.
#  Навмисно без чутливих полів: у листі немає ні алергій, ні
#  «кому не віддавати». Пошта — не найбезпечніше місце.
# ============================================================
def send_parent_mail(child_id, data):
    if not (SMTP_USER and SMTP_PASS):
        return

    # Звертаємось на ім'я. Раніше воно вгадувалося як друге слово ПІБ —
    # тепер поле окреме, гадати не треба. Відкат на друге слово лишений
    # для давніх заявок, де складових ще немає.
    who = data.get('parent_name', '').split()
    greeting = data.get('parent_first') or (who[1] if len(who) > 1 else (who[0] if who else ''))

    school = data.get('school', '')
    grade  = data.get('grade', '')
    line_school = '{}{}'.format(school, ', {} клас'.format(grade) if grade else '') if school else ''

    html = """<!doctype html><html><body style="margin:0;background:#F5F6F7;
padding:24px;font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif">
<div style="max-width:560px;margin:0 auto;background:#fff;border-radius:12px;
overflow:hidden;border:1px solid #E4E7EC">

  <div style="background:#2F6F4E;color:#fff;padding:22px 26px">
    <div style="font-size:13px;opacity:.85">ГО «Берегиня» · Освітній простір Покров</div>
    <div style="font-size:20px;font-weight:700;margin-top:6px">Заявку прийнято</div>
  </div>

  <div style="padding:24px 26px;color:#101828;font-size:16px;line-height:1.55">
    <p style="margin:0 0 16px">Доброго дня{greet}!</p>

    <p style="margin:0 0 16px">Ми отримали заявку на <b>{child}</b>.{school}</p>

    <p style="margin:0 0 8px"><b>Що далі</b></p>
    <p style="margin:0 0 16px">Ми зателефонуємо протягом двох робочих днів, щоб
    узгодити дні та години відвідування й відповісти на ваші запитання.</p>

    <p style="margin:0 0 16px">При першій зустрічі попросимо заповнити паперову
    картку дитини — там детальніше про здоров'я та перелік осіб, які можуть її
    забирати. Це займе близько п'яти хвилин.</p>

    <p style="margin:0 0 6px"><b>Якщо треба щось змінити</b></p>
    <p style="margin:0 0 20px">Просто зателефонуйте:<br>
      <a href="{h1}" style="color:#2F6F4E;font-weight:600;text-decoration:none">{p1}</a><br>
      <a href="{h2}" style="color:#2F6F4E;font-weight:600;text-decoration:none">{p2}</a>
    </p>

    <p style="margin:0;color:#667085;font-size:14px">
      Дані дитини ми обробляємо відповідно до
      <a href="{url}/dani.html" style="color:#475467">повідомлення про обробку
      персональних даних</a>.
    </p>
  </div>
</div>

<div style="max-width:560px;margin:14px auto 0;color:#98A2B3;font-size:12px;text-align:center">
  Цей лист надіслано автоматично, відповідати на нього не потрібно.
</div>
</body></html>""".format(
        greet=(', ' + esc(greeting)) if greeting else '',
        child=esc(data.get('child_name', '')),
        school=(' Заклад: ' + esc(line_school) + '.') if line_school else '',
        p1=PHONE_1, h1='tel:' + PHONE_1.replace(' ', ''),
        p2=PHONE_2, h2='tel:' + PHONE_2.replace(' ', ''),
        url=SITE_URL)

    msg = MIMEText(html, 'html', 'utf-8')
    subject = ('Заявку на підготовку до НМТ прийнято'
               if program_for(data.get('grade')) == 'НМТ'
               else 'Заявку до групи подовженого дня прийнято')
    msg['Subject'] = Header(subject, 'utf-8')
    msg['From']    = formataddr((str(Header('ГО «Берегиня»', 'utf-8')), MAIL_FROM))
    msg['To']      = data['parent_email']

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.sendmail(MAIL_FROM, [data['parent_email']], msg.as_string())
    app.logger.info('Лист батькам надіслано: заявка %s', child_id)


def esc(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;')
                  .replace('>', '&gt;').replace('"', '&quot;'))

# ============================================================
#  ВХІД
# ============================================================
def attempts_recently(ip):
    """Невдалі спроби ПЕРСОНАЛУ з цієї адреси за останні 15 хвилин.

    Умова login IS NULL тут не косметична: учнівські спроби пишуться в ту саму
    таблицю, але з логіном. Без цієї умови клас, який помиляється паролем із
    шкільної мережі, замкнув би адміністратору вхід у власну адмінку."""
    since = (datetime.now(timezone.utc).astimezone() - timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
    row = db().execute("SELECT COUNT(*) c FROM login_attempts "
                       "WHERE ip=? AND ts>? AND ok=0 AND (login IS NULL OR login='')",
                       (ip, since)).fetchone()
    return row['c']

@app.post('/api/login')
def login():
    ip = client_ip()
    if attempts_recently(ip) >= 7:
        return jsonify(ok=False, error='Забагато спроб. Спробуйте за 15 хвилин.'), 429

    login_ = clean(request.form.get('login'), 60)
    passwd = request.form.get('password') or ''
    row = db().execute('SELECT * FROM admins WHERE login=?', (login_,)).fetchone()
    ok = bool(row) and check_password_hash(row['pass_hash'], passwd)

    db().execute('INSERT INTO login_attempts (ip, ts, ok) VALUES (?,?,?)', (ip, now(), 1 if ok else 0))
    db().commit()

    if not ok:
        return jsonify(ok=False, error='Невірний логін або пароль'), 401

    session.clear()
    session.permanent = True
    session['kind']  = 'staff'          # див. gate(): учень цієї ознаки не отримує ніколи
    session['login'] = row['login']
    session['role']  = row['role']
    session['name']  = row['full_name'] or row['login']
    db().execute('UPDATE admins SET last_login=? WHERE id=?', (now(), row['id']))
    db().commit()
    log('вхід')
    return jsonify(ok=True, redirect='/admin')

@app.get('/api/logout')
@app.post('/api/logout')
def logout():
    if session.get('login'):
        log('вихід')
    session.clear()
    return redirect('/enter.html')

def require_login():
    if session.get('kind') != 'staff' or not session.get('login'):
        return redirect('/enter.html')
    return None


# ============================================================
#  ЄДИНА ЗАСТАВА НА ВХОДІ
#
#  Перевіряти роль у кожному маршруті окремо працює рівно доти,
#  доки хтось не забуде це зробити в одному новому маршруті —
#  і цього одного разу досить. Тому доступ вирішується тут, за
#  префіксом шляху, а перевірки в самих маршрутах лишаються як
#  другий рубіж.
#
#  /admin/*   — тільки персонал (kind='staff')
#  /kabinet/* — тільки учень   (kind='pupil')
#  Сесія при кожному вході очищається, тож учень фізично не може
#  мати ознаку staff, навіть якщо зайде відразу після педагога.
# ============================================================
@app.before_request
def gate():
    p = request.path
    if p.startswith('/admin'):
        if session.get('kind') != 'staff':
            return redirect('/enter.html')
    elif p.startswith('/kabinet'):
        if session.get('kind') != 'pupil':
            return redirect('/uchen.html')


# ============================================================
#  ДОСТУП УЧНІВ ДО НАВЧАЛЬНИХ МАТЕРІАЛІВ
#
#  Логін і пароль дитини — теж персональні дані, тому:
#   * пароль зберігається лише хешем, як і в персоналу;
#   * пароль показується один раз, у момент видачі;
#   * доступ прив'язаний до навчального року й гасне сам;
#   * спроби входу рахуються за логіном, а не за IP — інакше одна
#     дитина, що сім разів помилилася, замкнула б увесь клас,
#     який заходить із однієї шкільної адреси.
# ============================================================
UA2LAT = {
    'а':'a','б':'b','в':'v','г':'h','ґ':'g','д':'d','е':'e','є':'ie','ж':'zh',
    'з':'z','и':'y','і':'i','ї':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
    'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts',
    'ч':'ch','ш':'sh','щ':'shch','ь':'','ю':'iu','я':'ia','\'':'','’':'',
}

def translit(text):
    out = []
    for ch in (text or '').lower():
        if ch in UA2LAT:
            out.append(UA2LAT[ch])
        elif ch.isascii() and ch.isalnum():
            out.append(ch)
    return ''.join(out)


def make_login(last, cid, full=''):
    """Логін = прізвище латиницею + номер заявки.

    Номер потрібен не для краси: Ковальських у групі буває двоє.
    Читається з паперового талона без помилок і не змінюється ніколи.

    Раніше прізвище вгадувалося як перше слово повного імені — і варто
    було батькам написати «Іван Петренко», як логін ставав ivan12.
    Тепер береться окрема колонка. Відкат на перше слово лишений
    для заявок, поданих до розділення полів."""
    surname = (last or '').strip() or ((full or '').split() or [''])[0]
    return '{}{}'.format(translit(surname)[:14] or 'uchen', cid)


# без 0/o, 1/l/i — саме на них діти й помиляються, переписуючи з талона
PASS_ALPHABET = 'abcdefghjkmnpqrstuvwxyz23456789'

def gen_password():
    part = lambda: ''.join(secrets.choice(PASS_ALPHABET) for _ in range(4))
    return part() + '-' + part()


def norm_pass(v):
    """Дефіс у паролі — лише щоб його було легше переписати з аркуша.
    Хто його не набрав — не має отримати «невірний пароль»."""
    return re.sub(r'[^a-z0-9]', '', (v or '').strip().lower())


def pupil_attempts(login_):
    since = (datetime.now(timezone.utc).astimezone() - timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
    row = db().execute(
        'SELECT COUNT(*) c FROM login_attempts WHERE login=? AND ts>? AND ok=0',
        (login_, since)).fetchone()
    return row['c']


@app.post('/api/pupil-login')
def pupil_login():
    login_ = re.sub(r'[^a-z0-9]', '', clean(request.form.get('login'), 40).lower())
    passwd = norm_pass(request.form.get('password'))

    if not login_ or not passwd:
        return jsonify(ok=False, error='Введіть логін і пароль'), 400

    # ліміт за логіном: клас із однієї шкільної IP не має блокувати сам себе
    if pupil_attempts(login_) >= 10:
        return jsonify(ok=False, error='Забагато спроб. Спробуйте за 15 хвилин '
                                       'або підійдіть до викладача.'), 429

    row = db().execute(
        'SELECT id, child_name, grade, program, school_year, pass_hash '
        'FROM children WHERE login=?', (login_,)).fetchone()
    ok = bool(row) and bool(row['pass_hash']) and check_password_hash(row['pass_hash'], passwd)

    db().execute('INSERT INTO login_attempts (ip, ts, ok, login) VALUES (?,?,?,?)',
                 (client_ip(), now(), 1 if ok else 0, login_))
    db().commit()

    if not ok:
        return jsonify(ok=False, error='Невірний логін або пароль'), 401

    # доступ живе один навчальний рік і гасне сам — інакше через три роки
    # в базі були б живі логіни людей, які давно випустилися
    if (row['school_year'] or '') != school_year():
        return jsonify(ok=False, error='Доступ на цей навчальний рік не активний. '
                                       'Зверніться до викладача.'), 403

    session.clear()
    session.permanent = True
    session['kind']  = 'pupil'
    session['pupil'] = row['id']
    session['name']  = row['child_name']
    db().execute('UPDATE children SET last_seen=? WHERE id=?', (now(), row['id']))
    db().commit()
    log('вхід учня', row['id'], who='учень ' + login_)
    return jsonify(ok=True, redirect='/kabinet')


@app.get('/api/pupil-logout')
@app.post('/api/pupil-logout')
def pupil_logout():
    session.clear()
    return redirect('/uchen.html')


@app.get('/kabinet')
def kabinet():
    row = db().execute(
        'SELECT id, child_name, grade, program, school_year FROM children WHERE id=?',
        (session.get('pupil'),)).fetchone()
    if not row:                       # дитину видалили, поки сесія жила
        session.clear()
        return redirect('/uchen.html')
    # Рік показуємо той, що записаний у самій заявці, а не порахований
    # від сьогоднішньої дати: так на екрані завжди те, що в базі.
    return render_template('kabinet.html', c=row)


@app.post('/admin/pupil-password')
def admin_pupil_password():
    """Видає учневі логін і новий пароль. Пароль повертається один раз."""
    r = require_login()
    if r: return r

    ids = [x for x in (request.form.get('ids') or '').split(',') if x.strip().isdigit()]
    if len(ids) != 1:
        return jsonify(ok=False, error='Вибраний не один учень'), 400

    cid = int(ids[0])
    row = db().execute('SELECT id, child_name, child_last, grade, login, school_year '
                       'FROM children WHERE id=?', (cid,)).fetchone()
    if not row:
        return jsonify(ok=False, error='Такого учня немає'), 404

    login_ = row['login'] or make_login(row['child_last'], row['id'], row['child_name'])
    passwd = gen_password()
    db().execute('UPDATE children SET login=?, pass_hash=?, pass_set_at=?, pass_by=? WHERE id=?',
                 (login_, generate_password_hash(norm_pass(passwd)), now(),
                  session.get('login', '—'), cid))
    db().commit()
    log('видано пароль учню', cid)
    # Заявка минулих років: пароль видати можна, але ввійти з ним не вийде —
    # доступ прив'язаний до навчального року. Краще сказати про це одразу,
    # ніж потім розбиратися, чому дитина «не заходить».
    active = (row['school_year'] or '') == school_year()
    return jsonify(ok=True, id=cid, name=row['child_name'],
                   grade=row['grade'], login=login_, password=passwd,
                   renewed=bool(row['login']), active=active,
                   year=row['school_year'] or '—')

# ============================================================
#  АДМІНКА
# ============================================================
@app.get('/admin')
def admin_list():
    r = require_login()
    if r: return r

    q      = clean(request.args.get('q'), 60)
    status = clean(request.args.get('status'), 30)
    grade  = clean(request.args.get('grade'), 4)
    prog   = clean(request.args.get('program'), 8)
    year   = clean(request.args.get('year'), 9)

    sql, args = ('SELECT c.*, s.has_allergy FROM children c '
                 'LEFT JOIN sensitive s ON s.child_id=c.id WHERE 1=1'), []
    if prog in PROGRAMS:
        sql += ' AND c.program=?'
        args.append(prog)
    if year:
        sql += ' AND c.school_year=?'
        args.append(year)
    if q:
        sql += (' AND (lower_uk(c.child_name) LIKE ? OR lower_uk(c.parent_name) LIKE ?'
                ' OR c.parent_phone LIKE ?)')
        args += ['%' + q.lower() + '%'] * 3
    if status:
        sql += ' AND c.status=?'
        args.append(status)
    if grade:
        sql += ' AND c.grade=?'
        args.append(grade)
    # За прізвищем, а не за номером: список для друку має бути за абеткою.
    # Тепер прізвище — окрема колонка, тож абетка справді за прізвищем,
    # а не за тим словом, яке батько написав першим. COALESCE — для
    # заявок, поданих до розділення полів: там складових немає.
    sql += (' ORDER BY COALESCE(NULLIF(c.child_last,\'\'), c.child_name) COLLATE NOCASE,'
            ' c.child_first COLLATE NOCASE')

    rows = db().execute(sql, args).fetchall()
    counts = {r['status']: r['n'] for r in
              db().execute('SELECT status, COUNT(*) n FROM children GROUP BY status')}
    # класи показуємо лише ті, що є в межах обраного напряму —
    # інакше у фільтрі ГПД висіли б 9–11, яких там ніколи не буде
    gsql = 'SELECT DISTINCT grade FROM children WHERE grade<>""'
    gargs = []
    if prog in PROGRAMS:
        gsql += ' AND program=?'
        gargs.append(prog)
    grades = [r['grade'] for r in db().execute(
        gsql + ' ORDER BY CAST(grade AS INTEGER)', gargs)]

    progs = {r['program']: r['n'] for r in db().execute(
        'SELECT program, COUNT(*) n FROM children GROUP BY program')}

    # Перемикач років показуємо лише коли років справді кілька —
    # у перший рік роботи він був би зайвим елементом на екрані.
    years = [r['school_year'] for r in db().execute(
        'SELECT DISTINCT school_year FROM children '
        'WHERE school_year IS NOT NULL AND school_year<>"" ORDER BY school_year DESC')]

    log('перегляд списку')
    return render_template('admin.html', rows=rows, q=q, status=status, grade=grade,
                           program=prog, programs=PROGRAMS, prog_counts=progs,
                           year=year, years=years, this_year=school_year(),
                           grades=grades, counts=counts, total=sum(counts.values()),
                           monday=_monday(), keep=BACKUP_KEEP, statuses=STATUSES,
                           role=session.get('role'), who=session.get('name'))


def _monday(d=None):
    """Найближчий понеділок, від якого зручно починати тижневий список."""
    d = d or datetime.now().date()
    return (d - timedelta(days=d.weekday())).isoformat()


# ============================================================
#  СПИСОК НА ДРУК
#  Тільки те, що потрібно педагогу на аркуші: діти, контакти
#  і п'ять порожніх колонок під дати. Жодних статусів,
#  номерів заявок та іншої службової інформації.
# ============================================================
MONTHS = ['січня','лютого','березня','квітня','травня','червня',
          'липня','серпня','вересня','жовтня','листопада','грудня']

@app.get('/admin/spysok')
def admin_spysok():
    r = require_login()
    if r: return r

    raw = clean(request.args.get('ids'), 4000)
    ids = [int(x) for x in raw.split(',') if x.strip().isdigit()][:200]
    if not ids:
        return redirect('/admin')

    ph = ','.join('?' * len(ids))
    rows = [dict(x) for x in db().execute(
        'SELECT c.id, c.child_name, c.child_dob, c.grade, c.school, c.program, c.student_phone, '
        'c.parent_name, c.parent_phone, c.contact2_name, c.contact2_phone, n.subjects '
        'FROM children c LEFT JOIN nmt n ON n.child_id = c.id '
        'WHERE c.id IN ({}) '
        'ORDER BY CAST(c.grade AS INTEGER), '
        "COALESCE(NULLIF(c.child_last,''), c.child_name) COLLATE NOCASE".format(ph), ids)]

    pickups = {}
    for p in db().execute(
            'SELECT child_id, name, phone, relation FROM pickup_persons '
            'WHERE child_id IN ({}) ORDER BY child_id, ord'.format(ph), ids):
        pickups.setdefault(p['child_id'], []).append(dict(p))
    for row in rows:
        row['pickup'] = pickups.get(row['id'], [])

    # п'ять робочих днів від обраної дати; вихідні пропускаємо —
    # група працює лише з понеділка до п'ятниці
    try:
        start = datetime.strptime(clean(request.args.get('from'), 10), '%Y-%m-%d').date()
    except ValueError:
        start = datetime.strptime(_monday(), '%Y-%m-%d').date()

    days, d = [], start
    while len(days) < 5:
        if d.weekday() < 5:
            days.append(d)
        d += timedelta(days=1)

    # у заголовку — перелік класів, які потрапили до списку:
    # «3 клас», «3 і 4 клас» або «2, 3, 7 клас»
    gset = sorted({row['grade'] for row in rows if row['grade']},
                  key=lambda x: int(x) if x.isdigit() else 99)
    if not gset:
        grades_label = ''
    elif len(gset) == 1:
        grades_label = '{} клас'.format(gset[0])
    elif len(gset) == 2:
        grades_label = '{} і {} клас'.format(*gset)
    else:
        grades_label = '{} клас'.format(', '.join(gset))

    # якщо всі в списку одного напряму — пишемо його в заголовку аркуша
    pset = {row['program'] for row in rows if row.get('program')}
    program_label = list(pset)[0] if len(pset) == 1 else ''

    short = request.args.get('mode') == 'short'
    # «Без дат»: заняття переносять, і не завжди по порядку. Тоді зручніше
    # надрукувати порожні колонки й вписати числа ручкою просто на аркуші.
    nodates = request.args.get('nodates') == '1'

    other = '/admin/spysok?ids={}&from={}{}'.format(
        raw, request.args.get('from', ''), '&nodates=1' if nodates else '')
    if not short:
        other = other.replace('?', '?mode=short&', 1)

    today = datetime.now().date()
    log('список на друк ({}{}): {} дітей'.format(
        'короткий' if short else 'повний', ', без дат' if nodates else '', len(rows)))
    return render_template(
        'spysok.html',
        rows=rows,
        short=short,
        other_url=other,
        days=([''] * 5 if nodates
              else ['{:02d}.{:02d}'.format(x.day, x.month) for x in days]),
        nodates=nodates,
        period='{} – {} {}'.format(days[0].day, days[-1].day, MONTHS[days[-1].month - 1]),
        grades_label=grades_label,
        program_label=program_label,
        # Повний список НМТ має інший набір колонок. Умова program_label
        # означає «усі в списку одного напряму» — змішаних списків адмінка
        # не дає зібрати, але посилання можна набрати й руками.
        nmt_full=(not short and program_label == 'НМТ'),
        mixed=(len(pset) > 1),
        today='{} {} {}'.format(today.day, MONTHS[today.month - 1], today.year),
        role=session.get('role'), who=session.get('name'))

@app.get('/admin/<int:cid>')
def admin_child(cid):
    r = require_login()
    if r: return r

    child = db().execute('SELECT * FROM children WHERE id=?', (cid,)).fetchone()
    if not child:
        abort(404)
    sens   = db().execute('SELECT * FROM sensitive WHERE child_id=?', (cid,)).fetchone()
    pickup = db().execute('SELECT * FROM pickup_persons WHERE child_id=? ORDER BY ord', (cid,)).fetchall()
    nmt    = db().execute('SELECT * FROM nmt WHERE child_id=?', (cid,)).fetchone()

    log('перегляд заявки', cid)
    return render_template('child.html', c=child, s=sens, pickup=pickup, nmt=nmt,
                           label=LABEL, statuses=STATUSES,
                           role=session.get('role'), who=session.get('name'))

@app.post('/admin/<int:cid>/status')
def admin_status(cid):
    r = require_login()
    if r: return r
    st = clean(request.form.get('status'), 30)
    if st in STATUSES:
        db().execute('UPDATE children SET status=? WHERE id=?', (st, cid))
        db().commit()
        log('статус → ' + st, cid)
    return redirect('/admin/{}'.format(cid))

@app.post('/admin/<int:cid>/note')
def admin_note(cid):
    r = require_login()
    if r: return r
    db().execute('UPDATE children SET note_admin=? WHERE id=?',
                 (clean(request.form.get('note_admin')), cid))
    db().commit()
    log('коментар', cid)
    return redirect('/admin/{}'.format(cid))

# ============================================================
#  ВИДАЛЕННЯ ЗАЯВОК
#
#  Потрібне не для порядку, а для сміття: посилання на анкету
#  ходить по руках учнів, і рано чи пізно хтось напише туди
#  дурницю замість прізвища.
#
#  Видаляємо назовсім, разом із сирою копією: лишати в базі те,
#  що людина видалила саме через його зміст, безглуздо. Шлях
#  повернення — копія бази, тому кнопка «Зробити копію бази»
#  стоїть поруч.
# ============================================================
@app.post('/admin/status-bulk')
def admin_status_bulk():
    """Ставить один статус усім позначеним. Дія зворотна, тому окремого
    підтвердження на сервері немає — воно є у браузері."""
    r = require_login()
    if r: return r

    st = clean(request.form.get('status'), 30)
    if st not in STATUSES:
        return jsonify(ok=False, error='Невідомий статус'), 400

    ids = [int(x) for x in (request.form.get('ids') or '').split(',')
           if x.strip().isdigit()][:500]
    if not ids:
        return jsonify(ok=False, error='Не вибрано жодного учня'), 400

    ph = ','.join('?' * len(ids))
    conn = db()
    # рахуємо лише тих, кому статус справді змінюється, — щоб у відповіді
    # не було «змінено 12», коли 10 із них уже мали цей статус
    n = conn.execute(
        'SELECT COUNT(*) c FROM children WHERE id IN ({}) AND status<>?'.format(ph),
        ids + [st]).fetchone()['c']
    conn.execute('UPDATE children SET status=? WHERE id IN ({})'.format(ph), [st] + ids)
    conn.commit()
    log('статус → {} ({} заявок)'.format(st, n))
    return jsonify(ok=True, status=st, changed=n, total=len(ids))


@app.post('/admin/delete')
def admin_delete():
    r = require_login()
    if r: return r
    if session.get('role') != 'admin':
        return jsonify(ok=False, error='Видаляти заявки може лише адміністратор'), 403

    ids = [int(x) for x in (request.form.get('ids') or '').split(',')
           if x.strip().isdigit()][:200]
    if not ids:
        return jsonify(ok=False, error='Не вибрано жодного учня'), 400

    conn = db()
    ph = ','.join('?' * len(ids))
    rows = conn.execute(
        'SELECT id, child_name FROM children WHERE id IN ({})'.format(ph), ids).fetchall()
    if not rows:
        return jsonify(ok=False, error='Таких заявок уже немає'), 404

    try:
        # сирі копії зовнішнім ключем не звʼязані — прибираємо самі
        conn.execute('DELETE FROM raw_submissions WHERE child_id IN ({})'.format(ph), ids)
        # решту (алергії, НМТ, хто забирає) забере ON DELETE CASCADE:
        # PRAGMA foreign_keys увімкнено для зʼєднання в db()
        conn.execute('DELETE FROM children WHERE id IN ({})'.format(ph), ids)
        _cleanup_orphans(conn)          # запобіжник, якщо каскад раптом не спрацював

        # Якщо не лишилось нічого — скидаємо нумерацію, щоб наступна заявка
        # була № 1. Поки в базі є хоч один рядок, номери не переспользовуються:
        # інакше лист «заявка № 5» через місяць указував би на іншу дитину.
        empty = conn.execute('SELECT COUNT(*) c FROM children').fetchone()['c'] == 0
        if empty:
            conn.execute("DELETE FROM sqlite_sequence WHERE name='children'")
        conn.commit()
    except Exception as e:
        conn.rollback()
        app.logger.error('Заявки не видалено: %s', e)
        return jsonify(ok=False, error='Не вдалося видалити: {}'.format(e)), 500

    # імʼя пишемо в журнал разом із видаленням: самого рядка вже немає,
    # і без цього в журналі лишився б тільки номер
    for row in rows:
        log('видалено заявку: {}'.format(row['child_name']), row['id'])

    return jsonify(ok=True, deleted=len(rows), empty=empty,
                   names=[row['child_name'] for row in rows])


# ============================================================
#  ВИВАНТАЖЕННЯ ВСІЄЇ БАЗИ НА ПОШТУ
#
#  Два вкладення, бо вони для різного:
#    * .xlsx (або .zip з CSV) — щоб відкрити й редагувати як завгодно;
#    * .db — точна резервна копія, з якої можна підняти сайт із нуля.
#
#  Лист містить УСІ персональні дані всіх дітей, включно з даними про
#  здоров'я. Тому: тільки роль admin, тільки на адресу з .env (ніколи
#  на адресу з форми) і обов'язковий запис у журнал доступу.
# ============================================================
TABLES_FOR_EXPORT = ('children', 'sensitive', 'nmt', 'pickup_persons')

SHEET_TITLE = {'children': 'Діти', 'sensitive': 'Особливості',
               'nmt': 'Довузівська', 'pickup_persons': 'Хто забирає'}

# Порядок сортування задаємо явно: у sensitive і nmt первинний ключ —
# child_id, колонки id там немає, і спільне «ORDER BY id» мовчки
# викидало б обидві таблиці з вивантаження.
ORDER_BY = {'children': 'id', 'sensitive': 'child_id', 'nmt': 'child_id',
            'pickup_persons': 'child_id, ord'}

# Хеш пароля учня в таблиці не потрібен нікому, а лист із ним ходить поштою
# й лежить у скриньці роками. Логін лишаємо: по ньому видно, кому вже видано
# доступ. Сам хеш є в резервній копії .db — цього досить.
EXPORT_SKIP = {'pass_hash'}


def _table_rows(conn, table):
    cur = conn.execute('SELECT * FROM {} ORDER BY {}'.format(table, ORDER_BY[table]))
    cols = [d[0] for d in cur.description]
    keep = [i for i, c in enumerate(cols) if c not in EXPORT_SKIP]
    head = [LABEL.get(cols[i], cols[i]) for i in keep]
    return head, [[('' if row[i] is None else row[i]) for i in keep] for row in cur.fetchall()]


def _build_workbook(conn):
    """Повертає (ім'я файла, байти, mime). Без openpyxl — ZIP із CSV,
    щоб кнопка працювала навіть на сервері, де бібліотеку не встановили."""
    stamp = datetime.now().strftime('%Y-%m-%d')
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        wb.remove(wb.active)
        for t in TABLES_FOR_EXPORT:
            try:
                head, rows = _table_rows(conn, t)
            except sqlite3.OperationalError:
                continue
            ws = wb.create_sheet(SHEET_TITLE.get(t, t))
            ws.append(head)
            for c in ws[1]:
                c.font = Font(bold=True)
            for r in rows:
                ws.append(r)
            ws.freeze_panes = 'A2'
            for i, name in enumerate(head, start=1):
                width = max(len(str(name)) + 2,
                            *(len(str(r[i - 1])) + 2 for r in rows[:200])) if rows else len(name) + 2
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width, 46)
        buf = io.BytesIO()
        wb.save(buf)
        return ('beregynia-{}.xlsx'.format(stamp), buf.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
            for t in TABLES_FOR_EXPORT:
                try:
                    head, rows = _table_rows(conn, t)
                except sqlite3.OperationalError:
                    continue
                sio = io.StringIO()
                sio.write('\ufeff')                 # BOM, інакше Excel покаже кракозябри
                w = csv.writer(sio, delimiter=';')
                w.writerow(head)
                w.writerows(rows)
                z.writestr('{}.csv'.format(SHEET_TITLE.get(t, t)), sio.getvalue())
        return ('beregynia-{}.zip'.format(stamp), buf.getvalue(), 'application/zip')


def _human(n):
    return '{:.1f} МБ'.format(n / 1048576.0) if n >= 1048576 else '{:.0f} КБ'.format(n / 1024.0)


def _write_backup():
    """Знімає копію бази у BACKUP_DIR і прибирає найстаріші.

    sqlite3.backup() — не те саме, що cp: він знімає узгоджений стан навіть
    тоді, коли хтось саме зараз пише в базу. Просте копіювання файла в режимі
    WAL дало б биту копію, і виявилося б це вже під час відновлення.

    Імена файлів мають вигляд beregynia-2026-09-14-2130.db, тому звичайне
    сортування за назвою — це сортування за часом, і найстаріший завжди
    перший. Тримаємо останні BACKUP_KEEP штук: сервер не безрозмірний."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    # секунди в імені не для краси: без них дві копії поспіль (натиснули
    # «тільки на сервері», потім передумали й натиснули «на пошту») мовчки
    # перезаписали б одна одну
    path = os.path.join(BACKUP_DIR, 'beregynia-{}.db'.format(
        datetime.now().strftime('%Y-%m-%d-%H%M%S')))

    dst = sqlite3.connect(path)
    try:
        db().backup(dst)
    finally:
        dst.close()
    for extra in (path + '-wal', path + '-shm'):
        try: os.remove(extra)
        except OSError: pass

    files = sorted(f for f in os.listdir(BACKUP_DIR)
                   if f.startswith('beregynia-') and f.endswith('.db'))
    dropped = 0
    while len(files) > BACKUP_KEEP:
        try:
            os.remove(os.path.join(BACKUP_DIR, files.pop(0)))
            dropped += 1
        except OSError:
            break
    total = sum(os.path.getsize(os.path.join(BACKUP_DIR, f)) for f in files)
    return {'path': path, 'name': os.path.basename(path),
            'size': os.path.getsize(path), 'kept': len(files),
            'total': total, 'dropped': dropped}


@app.post('/admin/backup')
def admin_backup():
    """Копія бази. Спершу — завжди на сервер, і лише потім, за бажанням,
    на пошту. Порядок саме такий навмисно: якщо пошта не працює або лист
    ріжеться через розмір, копія на диску вже є."""
    r = require_login()
    if r: return r
    if session.get('role') != 'admin':
        return jsonify(ok=False, error='Копія бази доступна лише адміністратору'), 403

    want_mail = request.form.get('mail') == '1'

    try:
        bk = _write_backup()
    except Exception as e:
        app.logger.error('Копію бази не збережено: %s', e)
        return jsonify(ok=False, error='Не вдалося зберегти копію: {}'.format(e)), 500

    conn = db()
    n = conn.execute('SELECT COUNT(*) c FROM children').fetchone()['c']
    log('копія бази на сервері: {} ({} заявок)'.format(bk['name'], n))

    base = dict(ok=True, mailed=False, n=n, name=bk['name'],
                size=_human(bk['size']), kept=bk['kept'],
                total=_human(bk['total']), dir=BACKUP_DIR)
    if not want_mail:
        return jsonify(**base)

    if not (SMTP_USER and SMTP_PASS and ADMIN_MAIL):
        return jsonify(mail_error='Пошта не налаштована — див. .env', **base)

    fname, fbytes, fmime = _build_workbook(conn)

    snapshot, snap_note = None, ''
    try:
        with open(bk['path'], 'rb') as fh:
            snapshot = fh.read()
        if len(snapshot) > 20 * 1024 * 1024:      # поштові сервери ріжуть великі вкладення
            snap_note = ('Резервну копію бази не вкладено — вона більша за 20 МБ. '
                         'Копія збережена на сервері: {}'.format(bk['name']))
            snapshot = None
    except Exception as e:
        app.logger.error('Копію не прочитано для листа: %s', e)
        snap_note = 'Резервну копію бази не вкладено через помилку читання.'

    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    html = """<!doctype html><html><body style="margin:0;background:#F5F6F7;padding:24px;
font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif">
<div style="max-width:560px;margin:0 auto;background:#fff;border:1px solid #E4E7EC;border-radius:12px;overflow:hidden">
  <div style="background:#2F6F4E;color:#fff;padding:20px 24px">
    <div style="font-size:20px;font-weight:700">Копія бази · {stamp}</div>
  </div>
  <div style="padding:22px 24px;color:#101828;font-size:15px;line-height:1.6">
    <p style="margin:0 0 14px">У базі <b>{n}</b> заявок. Вкладено:</p>
    <p style="margin:0 0 6px"><b>{fname}</b> — усі дані таблицею, відкривається
       в Excel чи Google Таблицях, редагуйте як завгодно.</p>
    <p style="margin:0 0 16px">{snap}</p>
    <p style="margin:0;color:#B42318;font-size:14px">
      Вкладення містять персональні дані всіх дітей, зокрема про здоров'я.
      Не пересилайте цей лист далі й не зберігайте його у спільних теках.</p>
  </div>
</div></body></html>""".format(
        stamp=stamp, n=n, fname=esc(fname),
        snap=(esc(snap_note) if snap_note else
              '<b>{}</b> — точна резервна копія бази.'.format(esc(bk['name']))))

    msg = MIMEMultipart()
    msg['Subject'] = Header('Копія бази «Берегиня» · {}'.format(stamp), 'utf-8')
    msg['From']    = formataddr((str(Header('ГО «Берегиня»', 'utf-8')), MAIL_FROM))
    msg['To']      = ADMIN_MAIL
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    def attach(name, payload, subtype):
        part = MIMEApplication(payload, _subtype=subtype)
        part.add_header('Content-Disposition', 'attachment', filename=name)
        msg.attach(part)

    attach(fname, fbytes, fmime.split('/')[-1])
    if snapshot:
        attach(bk['name'], snapshot, 'octet-stream')

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as srv:
            srv.starttls()
            srv.login(SMTP_USER, SMTP_PASS)
            srv.sendmail(MAIL_FROM, [a.strip() for a in ADMIN_MAIL.split(',')], msg.as_string())
    except Exception as e:
        # Копія на сервері вже є — це не провал операції, а лише невдала пошта.
        app.logger.error('Копію бази не надіслано: %s', e)
        return jsonify(mail_error='Лист не надіслано: {}'.format(e), **base)

    log('копію бази надіслано на пошту ({} заявок)'.format(n))
    base.update(mailed=True, file=fname,
                to=ADMIN_MAIL.split(',')[0].strip(), note=snap_note)
    return jsonify(**base)


@app.get('/api/whoami')
def whoami():
    return jsonify(login=session.get('login'), role=session.get('role'))

# ============================================================
#  КОМАНДНИЙ РЯДОК
# ============================================================
def cli():
    import sys
    args = sys.argv[1:]
    init_db()

    if not args:
        print('База готова:', DB_PATH)
        print('Запуск для розробки:  DEV=1 python3 app.py run')
        print('Створити користувача: python3 app.py adduser <логін> <пароль> [admin|teacher]')
        return

    if args[0] == 'adduser':
        if len(args) < 3:
            print('python3 app.py adduser <логін> <пароль> [admin|teacher] ["Повне ім’я"]'); return
        login_, passwd = args[1], args[2]
        role = args[3] if len(args) > 3 else 'admin'
        name = args[4] if len(args) > 4 else login_
        if len(passwd) < 10:
            print('Пароль закороткий — щонайменше 10 символів.'); return
        conn = sqlite3.connect(DB_PATH)
        try:
            conn.execute('INSERT INTO admins (login, pass_hash, role, full_name, created_at) VALUES (?,?,?,?,?)',
                         (login_, generate_password_hash(passwd), role, name,
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            conn.commit()
            print('Створено:', login_, '/', role)
        except sqlite3.IntegrityError:
            print('Такий логін уже є.')
        conn.close()

    elif args[0] == 'passwd':
        if len(args) < 3:
            print('python3 app.py passwd <логін> <новий пароль>'); return
        conn = sqlite3.connect(DB_PATH)
        n = conn.execute('UPDATE admins SET pass_hash=? WHERE login=?',
                         (generate_password_hash(args[2]), args[1])).rowcount
        conn.commit(); conn.close()
        print('Оновлено' if n else 'Немає такого логіна')

    elif args[0] == 'run':
        app.run(host='127.0.0.1', port=int(os.environ.get('PORT', '8000')), debug=DEV)

    else:
        print('Невідома команда:', args[0])

if __name__ == '__main__':
    cli()
else:
    init_db()      # під gunicorn база теж має існувати
